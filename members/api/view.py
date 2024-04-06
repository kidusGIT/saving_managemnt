from django.shortcuts import redirect
from django.shortcuts import HttpResponse
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.generics import ListCreateAPIView
from django.contrib.auth.decorators import login_required
from datetime import datetime
from django.utils import timezone
from openpyxl import *
from openpyxl.styles import Font, Alignment

#  IMPORT MODULES HERE
from members.models import Member
from .serilizers import MemberSerilizer
from savings.serilizers import SavingTransactionSerilizer, CompulsoryTransactionSerilizer
from employees.models import Employee
from savings.models import CompulsorySaving, SavingAccount, CompulsoryTransaction, SavingTransaction

# CREATE MEMBER


@login_required
@api_view(['POST'])
def create_member(request):
    employee_id = request.user.employee_id
    employee = Employee.objects.get(employee_id=employee_id)

    if not employee.is_accountant:
        print('not accountant')
        return redirect('home-page')

    serilizer = MemberSerilizer(data=request.data)
    if serilizer.is_valid():
        serilizer.save()

    return Response(serilizer.data)

# UPDATE MEMBER


@api_view(['PUT'])
@login_required
def update_member(request, pk):
    employee_id = request.user.employee_id
    employee = Employee.objects.get(employee_id=employee_id)

    if not employee.is_accountant:
        return redirect('home-page')

    member = Member.objects.get(member_id=pk)
    serilizer = MemberSerilizer(instance=member, data=request.data)
    if serilizer.is_valid():
        serilizer.save()

    return Response(serilizer.data)

# MEMBER LIST


@api_view(['GET'])
@login_required
def members_list(request):
    employee_id = request.user.employee_id
    employee = Employee.objects.get(employee_id=employee_id)

    if employee.is_accountant or employee.is_credit_class:
        members = Member.objects.all().order_by('-created_date')
        serilizer = MemberSerilizer(members, many=True)

        return Response(serilizer.data)

    else:
        return redirect('home-page')


# SEARCH MEMBERS
class SearchMemberListView(ListCreateAPIView):
    queryset = Member.objects.all()
    serializer_class = MemberSerilizer
    filter_backends = [SearchFilter, OrderingFilter]

    search_fields = ['member_id', 'first_name',
                     'father_name', 'garnd_father_name']

# DELETE MEMBER


@api_view(['GET'])
@login_required
def delete_member(request, pk):
    print('deleted')
    employee_id = request.user.employee_id
    employee = Employee.objects.get(employee_id=employee_id)

    if not employee.is_accountant:
        return redirect('home-page')

    compulsory = CompulsorySaving.objects.get(member_id=pk)
    saving = SavingAccount.objects.get(member_id=pk, is_active=True)

    print('saving: ', saving.balance)
    print('compulsory: ', compulsory.balance)

    if saving.balance == 0.0 and compulsory.balance == 0.0:
        member = Member.objects.get(member_id=pk)
        member.delete()
        return Response({'delete': True})

    return Response({'delete': False})


@api_view(['GET'])
@login_required
def check_balance(request, pk):
    employee_id = request.user.employee_id
    employee = Employee.objects.get(employee_id=employee_id)

    if not employee.is_accountant:
        return redirect('home-page')

    compulsory = CompulsorySaving.objects.get(member_id=pk)
    saving = SavingAccount.objects.get(member_id=pk, is_active=True)

    if saving.balance == 0.0 and compulsory.balance == 0.0:
        return Response({'delete': True})

    return Response({'delete': False})

# MEMBER DETAIL


@api_view(['GET'])
@login_required
def get_member_detail(request, pk):

    employee_id = request.user.employee_id
    employee = Employee.objects.get(employee_id=employee_id)

    if not employee.is_accountant:
        return redirect('home-page')

    member = Member.objects.get(member_id=pk)
    serilizer = MemberSerilizer(member, many=False)

    return Response(serilizer.data)


report_for_other_saving = []
other_start_time = ''
other_end_time = ''
# COMPULSORY SAVING TABLE POPULATOR


@api_view(['POST'])
@login_required
def generate_other_report(request, check):

    start = request.POST.get('start')
    end = request.POST.get('end')

    global other_end_time
    global other_start_time

    other_start_time = get_date(start)
    other_end_time = get_date(end)

    report = SavingTransaction.objects.filter(
        created_date__range=(start, end)).select_related('account')
    serilizer = SavingTransactionSerilizer(report, many=True)

    global report_for_other_saving
    report_for_other_saving = report

    return Response(serilizer.data)

# OTHER SAVING EXCEL GENERATOR


def other_saving_excel(request):
    report = report_for_other_saving

    name = "members transaction.xlsx"
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + name + ''

    workbook = Workbook()
    worksheet = workbook.create_sheet(index=0, title='Transaction')

    worksheet.merge_cells('A1:K2')
    title = worksheet.cell(row=1, column=1)
    title.alignment = Alignment(horizontal='center', vertical='center')
    title.font = Font(bold=True, size=15)
    title.value = 'Transaction report for other type of saving from ' + \
        other_start_time + ' - ' + other_end_time

    columns = [
        'Member ID', 'Member Name', 'Father Name',
        'Grand father name', 'Email', 'Account No', 'Saving Type',
        'Balance', 'Withdraw', 'Deposit', 'Transaction Date'
    ]

    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['C'].width = 25
    worksheet.column_dimensions['D'].width = 25
    worksheet.column_dimensions['E'].width = 25
    worksheet.column_dimensions['F'].width = 15
    worksheet.column_dimensions['G'].width = 25
    worksheet.column_dimensions['K'].width = 25

    for column in range(len(columns)):
        header = worksheet.cell(row=3, column=column + 1)
        header.font = Font(bold=True)
        header.value = columns[column]

    row = 4
    for transaction in report:
        worksheet.cell(
            row=row, column=1).value = transaction.account.member.member_id
        worksheet.cell(
            row=row, column=2).value = transaction.account.member.first_name
        worksheet.cell(
            row=row, column=3).value = transaction.account.member.father_name
        worksheet.cell(
            row=row, column=4).value = transaction.account.member.garnd_father_name
        worksheet.cell(
            row=row, column=5).value = transaction.account.member.email
        worksheet.cell(
            row=row, column=6).value = transaction.account.account_id
        worksheet.cell(
            row=row, column=7).value = transaction.account.saving_type.account_name
        worksheet.cell(row=row, column=8).value = transaction.account.balance
        worksheet.cell(row=row, column=9).value = transaction.withdraw
        worksheet.cell(row=row, column=10).value = transaction.deposit
        worksheet.cell(row=row, column=11).value = get_date(
            transaction.created_date)

        row += 1

    workbook.save(response)

    return response
    # return HttpResponse('hello world')


report_for_compulsory_saving = []
compulsory_start_time = ''
compulsory_end_time = ''
# COMPULSORY SAVING REPORT TABLE POPULATOR


@api_view(['POST'])
@login_required
def generate_compulsory_report(request, check):

    start = request.POST.get('start')
    end = request.POST.get('end')

    global compulsory_end_time
    global compulsory_start_time

    compulsory_start_time = get_date(start)
    compulsory_end_time = get_date(end)

    # start_time = datetime.strptime(
    #     start, "%Y-%m-%dT%H:%M:%S").strftime('%Y-%m-%d %H:%M')

    print(get_date(start))

    report = CompulsoryTransaction.objects.filter(
        created_date__range=(start, end)).select_related('account')
    serilizer = CompulsoryTransactionSerilizer(report, many=True)

    global report_for_compulsory_saving
    report_for_compulsory_saving = report

    return Response(serilizer.data)

# COMPULSORY SAVING EXCEL GENERATOR


def compulsory_saving_excel(request):
    report = report_for_compulsory_saving

    name = "members transaction.xlsx"
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + name + ''

    workbook = Workbook()
    worksheet = workbook.create_sheet(index=0, title='Transaction')

    worksheet.merge_cells('A1:J2')
    title = worksheet.cell(row=1, column=1)
    title.alignment = Alignment(horizontal='center', vertical='center')
    title.font = Font(bold=True, size=15)
    title.value = 'Transaction report for compulsory type of saving from ' + \
        compulsory_start_time + ' - ' + compulsory_end_time

    columns = [
        'Member ID', 'Member Name', 'Father Name',
        'Grand father name', 'Email', 'Account No',
        'Balance', 'Withdraw', 'Deposit', 'Transaction Date'
    ]

    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['C'].width = 25
    worksheet.column_dimensions['D'].width = 25
    worksheet.column_dimensions['E'].width = 25
    worksheet.column_dimensions['F'].width = 15
    worksheet.column_dimensions['G'].width = 15
    worksheet.column_dimensions['H'].width = 15
    worksheet.column_dimensions['I'].width = 15
    worksheet.column_dimensions['J'].width = 25

    for column in range(len(columns)):
        header = worksheet.cell(row=3, column=column + 1)
        header.font = Font(bold=True)
        header.value = columns[column]

    row = 4

    for transaction in report:
        worksheet.cell(
            row=row, column=1).value = transaction.account.member.member_id
        worksheet.cell(
            row=row, column=2).value = transaction.account.member.first_name
        worksheet.cell(
            row=row, column=3).value = transaction.account.member.father_name
        worksheet.cell(
            row=row, column=4).value = transaction.account.member.garnd_father_name
        worksheet.cell(
            row=row, column=5).value = transaction.account.member.email
        worksheet.cell(
            row=row, column=6).value = transaction.account.account_id
        worksheet.cell(row=row, column=7).value = transaction.account.balance
        worksheet.cell(row=row, column=8).value = transaction.withdraw
        worksheet.cell(row=row, column=9).value = transaction.deposit
        worksheet.cell(row=row, column=10).value = get_date(
            transaction.created_date)

        row += 1

    workbook.save(response)

    return response


def get_date(date):

    return str(date)[:19]
