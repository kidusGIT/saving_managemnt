from django.shortcuts import render, redirect
from django.http import HttpResponseNotFound, HttpResponse
from django.contrib.auth.decorators import login_required
import openpyxl
from django.utils import timezone
from openpyxl.styles import Font, Alignment


# IMPORT USER DEFIND THINGS HERE
from .helpers import get_id
from .models import Member
from savings.models import SavingType, CompulsorySaving, SavingAccount
from savings.helpers import get_compulsory_id, get_savings_id
from employees.models import Employee


# MEMBERS LIST
@login_required
def member_list(request):
    employee_id = request.user.employee_id
    employee = Employee.objects.get(employee_id=employee_id)

    if not employee.is_accountant:
        return redirect('home-page')

    member_list = Member.objects.all().order_by('-created_date')
    return render(request, 'membersList.html', {'member_list': member_list})

# DEATIL MEMBER

# CREATE MEMBER


@login_required
def create_member(request):
    employee_id = request.user.employee_id
    employee = Employee.objects.get(employee_id=employee_id)

    if not employee.is_accountant:
        return redirect('home-page')

    member_id = get_id()
    compulsory_account = get_compulsory_id()
    saving_account = get_savings_id()

    savings = SavingType.objects.all()

    return render(request, 'add_member.html', {'id': member_id, 'compulsory_account': compulsory_account,
                                               'saving_account': saving_account, 'savings': savings})

# UPDATE MEMBER


@login_required
def update_member(request, pk):
    employee_id = request.user.employee_id
    employee = Employee.objects.get(employee_id=employee_id)

    if not employee.is_accountant:
        return redirect('home-page')

    member = Member.objects.get(member_id=pk)

    try:
        compulsoy_account = CompulsorySaving.objects.get(member=pk)
    except:
        return HttpResponseNotFound('No Memeber with ID ' + pk)

    saving_account = get_savings_id()

    savings = SavingAccount.objects.filter(member=pk).select_related(
        'saving_type').select_related('member')

    context = {
        'first_name': member.first_name,
        'father_name': member.father_name,
        'garnd_father_name': member.garnd_father_name,
        # address
        'city': member.city,
        'sub_city': member.sub_city,
        'woreda': member.woreda,
        'house_num': member.house_num,

        'email': member.email,
        'phone_num': member.phone_num,
        'age': member.age,
    }

    full_name = member.first_name + ' ' + member.father_name

    return render(request, 'member_info.html', {'member': context, 'id': pk, 'savings': savings,
                                                'saving_account': saving_account, 'compulsoy_account': compulsoy_account.account_id, 'full_name': full_name})

# ACCOUNTANT REPORTS


def account_report(request):
    now = timezone.now()
    return render(request, 'accountant_report.html', {'now': now})

# DELETE MEMBER


@login_required
def delete_member(request, pk):
    employee_id = request.user.employee_id
    employee = Employee.objects.get(employee_id=employee_id)

    if not employee.is_accountant:
        return redirect('home-page')

    return redirect('member-list')

# GENERATE EXCLE REPORT


@login_required
def export_xls(request):
    employee_id = request.user.employee_id
    employee = Employee.objects.get(employee_id=employee_id)

    if not employee.is_accountant:
        return redirect('home-page')

    members = Member.objects.all()
    name = "members " + members[0].first_name + ".xlsx"

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + name + ''

    wb = openpyxl.Workbook()

    sheet = wb.active
    wb.create_sheet(index=1, title='Sheet One')

    sheet.HeaderFooter.differentFirst = True
    sheet.HeaderFooter.firstHeader.left.text = 'Hello World'
    sheet.HeaderFooter.firstHeader.left.size = 18

    columns = ['First Name', 'Second Name', 'Last Name', 'Email']
    for c in range(len(columns)):
        pass
        # col = sheet.cell(row=1, column=c)
        # col.value = columns[c]

    sheet.merge_cells(start_column=1, end_column=6, start_row=1, end_row=2)

    title = sheet.cell(row=1, column=1)
    title.alignment = Alignment(horizontal='center', vertical='center')
    title.font = Font(bold=True, size=18)
    title.value = 'Hello World'

    row = 3
    for member in members:

        first_name = sheet.cell(row=row, column=1)
        # first_name.font = Font(bold=True, size=20)
        first_name.value = member.first_name

        second_name = sheet.cell(row=row, column=2)
        second_name.value = member.father_name

        last_name = sheet.cell(row=row, column=3)
        last_name.value = member.garnd_father_name

        email = sheet.cell(row=row, column=4)
        email.value = member.email

        row = row + 1

    wb.save(response)

    return response
