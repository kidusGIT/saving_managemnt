from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
import datetime
# from datetime import datetime
import xlwt
from openpyxl import *
from openpyxl.styles import Font, Alignment

# IMPORT USER BUILT MODULE HERE
from members.models import Member
from savings.models import CompulsorySaving, CompulsoryTransaction, SavingAccount, SavingType, SavingTransaction

# MEMBERS WHOLE TRANSACTION


@login_required
def get_members_transaction(request, pk):
    saving_transaction = SavingTransaction.objects.filter(
        account__member__member_id__exact=pk).select_related('account').order_by('created_date')
    compulsory_transaction = CompulsoryTransaction.objects.filter(
        account__member__member_id__exact=pk).select_related('account').order_by('account').order_by('created_date')

    comlsory = CompulsorySaving.objects.get(member_id=pk)
    account = SavingAccount.objects.get(member_id=pk, is_active=True)

    balance = comlsory.balance + account.balance

    full_name = comlsory.member.first_name + ' ' + comlsory.member.father_name

    return render(request, 'membersTransaction.html', {'saving_transaction': saving_transaction,
                                                       'compulsory_transaction': compulsory_transaction, 'id': pk, 'balance': balance, 'full_name': full_name})

# MEMBERS COMPULSORY MONTHLY SAVING DISPLAY


@login_required
def compulsory_monthly_saving(request):

    return render(request, 'monthlySaving.html')


@login_required
def generate_report(request, pk, check):

    if check == 0:
        saving_transaction = SavingTransaction.objects.filter(
            account__member__member_id__exact=pk).select_related('account').order_by('created_date')

        full_name = saving_transaction[0].account.member.first_name + \
            ' ' + saving_transaction[0].account.member.father_name
        name = "members " + full_name + ".xlsx"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=' + name + ''

        workbook = Workbook()
        worksheet = workbook.create_sheet(
            index=0, title=full_name + ' transaction')

        worksheet.merge_cells('A1:K2')
        title = worksheet.cell(row=1, column=1)
        title.alignment = Alignment(horizontal='center', vertical='center')
        title.font = Font(bold=True, size=15)
        title.value = full_name + ', full transaction'

        row = 4
        columns = ['Member ID', 'First Name', 'Father Name', 'Grand Father Name',
                   'Email', 'Account Type', 'Account ID', 'Deposit', 'Withdraw', 'Balance', 'Date']

        worksheet.column_dimensions['A'].width = 13
        worksheet.column_dimensions['B'].width = 25
        worksheet.column_dimensions['C'].width = 25
        worksheet.column_dimensions['D'].width = 25
        worksheet.column_dimensions['E'].width = 25
        worksheet.column_dimensions['F'].width = 25
        worksheet.column_dimensions['H'].width = 15
        worksheet.column_dimensions['I'].width = 15
        worksheet.column_dimensions['J'].width = 15
        worksheet.column_dimensions['G'].width = 15
        worksheet.column_dimensions['K'].width = 25

        for column in range(len(columns)):
            header = worksheet.cell(row=3, column=column + 1)
            header.font = Font(bold=True)
            header.value = columns[column]

        for transaction in saving_transaction:
            worksheet.cell(
                row=row, column=1).value = transaction.account.member.member_id
            worksheet.cell(
                row=row, column=2).value = transaction.account.member.first_name
            worksheet.cell(
                row=row, column=3).value = transaction.account.member.father_name
            worksheet.cell(
                row=row, column=4).value = transaction.account.member.garnd_father_name
            worksheet.cell(
                row=row, column=5).value = transaction.account.member.email
            worksheet.cell(
                row=row, column=6).value = transaction.account.saving_type.account_name
            worksheet.cell(
                row=row, column=7).value = transaction.account.account_id
            worksheet.cell(row=row, column=8).value = transaction.deposit
            worksheet.cell(row=row, column=9).value = transaction.withdraw
            worksheet.cell(row=row, column=10).value = transaction.balance
            worksheet.cell(row=row, column=11).value = get_date(
                transaction.created_date)

            row += 1

        workbook.save(response)

        return response

    elif check == 1:
        compulsory_transaction = CompulsoryTransaction.objects.filter(
            account__member__member_id__exact=pk).select_related('account').order_by('account').order_by('created_date')
        full_name = compulsory_transaction[0].account.member.first_name + ' ' + \
            compulsory_transaction[0].account.member.father_name
        # ' report on date ' + get_date(datetime.datetime.now())
        name = "members " + full_name + ".xlsx"

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=' + name + ''

        workbook = Workbook()
        worksheet = workbook.create_sheet(
            index=0, title=full_name + ' transaction')

        worksheet.merge_cells('A1:J2')
        title = worksheet.cell(row=1, column=1)
        title.alignment = Alignment(horizontal='center', vertical='center')
        title.font = Font(bold=True, size=15)
        title.value = full_name + ', full transaction'

        worksheet.column_dimensions['A'].width = 13
        worksheet.column_dimensions['B'].width = 25
        worksheet.column_dimensions['C'].width = 25
        worksheet.column_dimensions['D'].width = 25
        worksheet.column_dimensions['E'].width = 25
        worksheet.column_dimensions['F'].width = 18
        worksheet.column_dimensions['H'].width = 15
        worksheet.column_dimensions['I'].width = 15
        worksheet.column_dimensions['J'].width = 25
        worksheet.column_dimensions['G'].width = 15
        # worksheet.column_dimensions['K'].width = 25

        row = 4
        columns = ['Member ID', 'First Name', 'Father Name', 'Grand Father Name',
                   'Email', 'Account Number', 'Deposit', 'Withdraw', 'Balance', 'Date']

        for column in range(len(columns)):
            header = worksheet.cell(row=3, column=column + 1)
            header.font = Font(bold=True)
            header.value = columns[column]

        for transaction in compulsory_transaction:
            worksheet.cell(
                row=row, column=1).value = transaction.account.member.member_id
            worksheet.cell(
                row=row, column=2).value = transaction.account.member.first_name
            worksheet.cell(
                row=row, column=3).value = transaction.account.member.father_name
            worksheet.cell(
                row=row, column=4).value = transaction.account.member.garnd_father_name
            worksheet.cell(
                row=row, column=5).value = transaction.account.member.email
            worksheet.cell(
                row=row, column=6).value = transaction.account.account_id
            worksheet.cell(row=row, column=7).value = transaction.deposit
            worksheet.cell(row=row, column=8).value = transaction.withdraw
            worksheet.cell(row=row, column=9).value = transaction.balance
            worksheet.cell(row=row, column=10).value = get_date(
                transaction.created_date)

            row += 1

        workbook.save(response)

        return response


def get_date(date):

    return str(date)[:19]
